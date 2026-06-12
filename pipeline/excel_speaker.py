import re
import subprocess
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, filedialog, messagebox, colorchooser
from datetime import datetime, date, time
import json
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

class ExcelSpeakerApp:
    APP_SETTINGS_KEY = "__app_settings__"
    UI_SETTING_FIELDS = (
        ("voice_var", "voice"),
        ("rate_var", "rate"),
        ("eng_col_var", "eng_col"),
        ("eng_font_var", "eng_font"),
        ("eng_size_var", "eng_size"),
        ("eng_width_var", "eng_width"),
        ("eng_color_var", "eng_color"),
        ("chn_col_var", "chn_col"),
        ("chn_font_var", "chn_font"),
        ("chn_size_var", "chn_size"),
        ("chn_width_var", "chn_width"),
        ("chn_color_var", "chn_color"),
        ("page_size_var", "page_size"),
    )
    THEMES = {
        "dark": {
            "bg": "#0F1115",
            "panel": "#171A21",
            "panel_2": "#1D212B",
            "cell": "#141821",
            "header": "#202633",
            "grid": "#262D3A",
            "text": "#E7EFFF",
            "muted": "#9CA3AF",
            "subtle": "#C7CDD6",
            "accent": "#60A5FA",
            "accent_soft": "#1E3A5F",
            "button": "#232938",
            "button_hover": "#2B3345",
            "entry": "#11161F",
            "entry_border": "#374151",
        },
        "light": {
            "bg": "#F8FAFC",
            "panel": "#EEF2F7",
            "panel_2": "#E2E8F0",
            "cell": "#FFFFFF",
            "header": "#E8EEF8",
            "grid": "#D6DEE8",
            "text": "#111827",
            "muted": "#6B7280",
            "subtle": "#374151",
            "accent": "#2563EB",
            "accent_soft": "#DBEAFE",
            "button": "#FFFFFF",
            "button_hover": "#F3F4F6",
            "entry": "#FFFFFF",
            "entry_border": "#CBD5E1",
        },
    }

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Excel Speaker for English Practice")
        self.root.geometry("1300x800")
        
        self.style = ttk.Style(self.root)
        if "clam" in self.style.theme_names():
            self.style.theme_use("clam")

        app_dir = os.path.dirname(os.path.abspath(__file__))
        self.config_dir = os.path.join(app_dir, ".excel_speaker")
        self.config_file = os.path.join(self.config_dir, "excel_speaker_config.json")
        self.legacy_config_files = [
            os.path.join(
                os.path.expanduser("~"),
                "Library",
                "Application Support",
                "ExcelSpeaker",
                "excel_speaker_config.json",
            ),
            os.path.join(app_dir, "excel_speaker_config.json"),
            os.path.join(os.path.expanduser("~"), ".excel_speaker_config.json"),
        ]
        self.bookmark_data = self.load_config()
        self.color_buttons = {}
        self.theme_mode = self.get_app_settings().get("theme_mode", "dark")
        if self.theme_mode not in self.THEMES:
            self.theme_mode = "dark"
        self.theme = dict(self.THEMES[self.theme_mode])
        self.root.configure(bg=self.theme["bg"])
        self._configure_theme()

        # 核心数据
        self.workbook = None
        self.file_path = None
        self.say_proc = None
        self.current_page = 0
        self.total_pages = 1

        # 状态管理
        self.row_labels = {}
        self.cell_labels = {}
        self.current_highlight_row = None
        self.current_selected_col = None
        self.last_spoken_cell_info = None
        self.settings_visible = True
        self.single_click_job = None
        self.pending_click_args = None

        self.font_options = [
            "System Default", 
            "Avenir Next", "Trebuchet MS", "Optima", 
            "Palatino", "Georgia", 
            "Chalkboard SE", "Comic Sans MS"
        ]

        # === 变量定义 ===
        self.sheet_var = tk.StringVar()
        self.voice_var = tk.StringVar(value="Ava (Premium)")
        self.rate_var = tk.IntVar(value=165)
        
        # 英文列设置
        self.eng_col_var = tk.StringVar(value="A")
        self.eng_font_var = tk.StringVar(value="Avenir Next")
        self.eng_size_var = tk.IntVar(value=18)
        self.eng_width_var = tk.IntVar(value=500)
        self.eng_color_var = tk.StringVar(value="#111827") # 深灰黑
        
        # 中文列设置
        self.chn_col_var = tk.StringVar(value="B")
        self.chn_font_var = tk.StringVar(value="System Default")
        self.chn_size_var = tk.IntVar(value=16)
        self.chn_width_var = tk.IntVar(value=400)
        self.chn_color_var = tk.StringVar(value="#4B5563") # 较浅的灰色
        
        self.page_size_var = tk.IntVar(value=100)
        self.page_info_var = tk.StringVar(value="Page 0/0")
        self.status_var = tk.StringVar(value="Welcome! Open an Excel file to begin.")

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _configure_theme(self):
        t = self.theme

        self.style.configure(
            ".",
            background=t["bg"],
            foreground=t["text"],
            fieldbackground=t["entry"],
        )

        self.style.configure("TFrame", background=t["bg"])
        self.style.configure("TLabel", background=t["bg"], foreground=t["text"])

        self.style.configure(
            "TLabelframe",
            background=t["panel"],
            borderwidth=1,
            relief="solid",
            bordercolor=t["grid"],
        )
        self.style.configure(
            "TLabelframe.Label",
            background=t["panel"],
            foreground=t["text"],
        )

        self.style.configure(
            "TButton",
            background=t["button"],
            foreground=t["text"],
            borderwidth=1,
            focusthickness=0,
            padding=6,
            relief="flat",
        )
        self.style.map(
            "TButton",
            background=[("active", t["button_hover"]), ("pressed", t["accent_soft"])],
            foreground=[("disabled", t["muted"])],
        )
        self.style.configure("Picker.TButton", padding=6)

        self.style.configure(
            "TEntry",
            fieldbackground=t["entry"],
            foreground=t["text"],
            insertcolor=t["text"],
            bordercolor=t["entry_border"],
            lightcolor=t["entry_border"],
            darkcolor=t["entry_border"],
            padding=4,
        )

        self.style.configure(
            "TCombobox",
            fieldbackground=t["entry"],
            background=t["entry"],
            foreground=t["text"],
            arrowcolor=t["text"],
            bordercolor=t["entry_border"],
            lightcolor=t["entry_border"],
            darkcolor=t["entry_border"],
            padding=4,
        )
        self.style.map(
            "TCombobox",
            fieldbackground=[("readonly", t["entry"])],
            foreground=[("readonly", t["text"])],
            selectbackground=[("readonly", t["entry"])],
            selectforeground=[("readonly", t["text"])],
        )

        self.style.configure(
            "TSpinbox",
            fieldbackground=t["entry"],
            background=t["entry"],
            foreground=t["text"],
            arrowcolor=t["text"],
            bordercolor=t["entry_border"],
            lightcolor=t["entry_border"],
            darkcolor=t["entry_border"],
            padding=4,
        )

        self.style.configure(
            "Vertical.TScrollbar",
            background=t["panel_2"],
            troughcolor=t["bg"],
            bordercolor=t["bg"],
            arrowcolor=t["text"],
        )
        self.style.configure(
            "Horizontal.TScrollbar",
            background=t["panel_2"],
            troughcolor=t["bg"],
            bordercolor=t["bg"],
            arrowcolor=t["text"],
        )

    def get_app_settings(self):
        settings = self.bookmark_data.get(self.APP_SETTINGS_KEY)
        return settings if isinstance(settings, dict) else {}

    def save_app_settings(self, quiet=True):
        settings = self.bookmark_data.get(self.APP_SETTINGS_KEY)
        if not isinstance(settings, dict):
            settings = {}
            self.bookmark_data[self.APP_SETTINGS_KEY] = settings
        settings["theme_mode"] = self.theme_mode
        self.save_config(quiet=quiet)

    def load_config(self):
        for path in [self.config_file, *self.legacy_config_files]:
            if not path or not os.path.exists(path):
                continue
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if path != self.config_file:
                    self.bookmark_data = data
                    self.save_config(quiet=True)
                return data
            except Exception as e:
                print(f"Load config failed: {path} | {e}")
        return {}

    def save_config(self, quiet=False):
        try:
            os.makedirs(self.config_dir, exist_ok=True)
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(self.bookmark_data, f, ensure_ascii=False, indent=2)
            if not quiet:
                self.status_var.set(f"Config saved: {self.config_file}")
        except Exception as e:
            print(f"Save config failed: {self.config_file} | {e}")

    def normalize_file_key(self, path):
        if not path:
            return ""
        return os.path.normcase(os.path.realpath(os.path.abspath(path)))

    def _ensure_file_store(self):
        if not self.file_path:
            return None
        file_key = self.normalize_file_key(self.file_path)
        store = self.bookmark_data.get(file_key)
        if not isinstance(store, dict):
            store = {}
            self.bookmark_data[file_key] = store
        return store

    def get_file_store(self):
        if not self.file_path:
            return {}

        file_key = self.normalize_file_key(self.file_path)
        store = self.bookmark_data.get(file_key)
        if isinstance(store, dict):
            return store

        legacy_keys = [
            self.file_path,
            os.path.abspath(self.file_path),
            os.path.realpath(self.file_path),
        ]
        for legacy_key in legacy_keys:
            store = self.bookmark_data.get(legacy_key)
            if isinstance(store, dict):
                self.bookmark_data[file_key] = store
                self.save_config()
                return store

        return {}

    def get_default_col_index(self):
        try:
            return column_index_from_string((self.eng_col_var.get() or "A").upper())
        except Exception:
            return 1

    def get_saved_sheet_name(self):
        store = self.get_file_store()
        sheet_name = store.get("__last_sheet__")
        if isinstance(sheet_name, str) and self.workbook is not None and sheet_name in self.workbook.sheetnames:
            return sheet_name
        return None

    def get_saved_position(self):
        store = self.get_file_store()
        resume_store = store.get("__resume_position__")
        if not isinstance(resume_store, dict):
            return None

        pos = resume_store.get(self.sheet_var.get())
        if not isinstance(pos, dict):
            return None

        try:
            row = max(1, int(pos.get("row", 1)))
            col = max(1, int(pos.get("col", self.get_default_col_index())))
            return {"row": row, "col": col}
        except Exception:
            return None

    def save_resume_position(self):
        if not self.file_path or not self.sheet_var.get():
            return

        store = self._ensure_file_store()
        if store is None:
            return

        row = self.current_highlight_row
        if row is None:
            page_size = max(1, int(self.page_size_var.get()))
            row = self.current_page * page_size + 1

        col = self.current_selected_col or self.get_default_col_index()
        resume_store = store.get("__resume_position__")
        if not isinstance(resume_store, dict):
            resume_store = {}
            store["__resume_position__"] = resume_store

        store["__last_sheet__"] = self.sheet_var.get()
        resume_store[self.sheet_var.get()] = {
            "row": max(1, int(row)),
            "col": max(1, int(col)),
        }
        self.save_config(quiet=True)

    def _build_ui(self):
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(1, weight=1)

        # ==========================================
        # 1. 顶部操作栏
        # ==========================================
        top_bar = ttk.Frame(self.root, padding=8)
        top_bar.grid(row=0, column=0, columnspan=2, sticky="ew")
        
        # 侧边栏开关
        ttk.Button(top_bar, text="⚙️ Settings", command=self.toggle_settings).pack(side="left", padx=(0, 15))
        self.theme_btn = ttk.Button(top_bar, command=self.toggle_theme)
        self.theme_btn.pack(side="left", padx=(0, 8))
        self._update_theme_button_text()
        
        ttk.Button(top_bar, text="📂 Open Excel", command=self.open_file).pack(side="left", padx=(0, 5))
        # ttk.Button(top_bar, text="🔄 Reload", command=self.load_current_sheet).pack(side="left", padx=5)
        ttk.Button(top_bar, text="⏹ Stop Speech", command=self.stop_speaking).pack(side="left", padx=5)

        ttk.Label(top_bar, text="Sheet:").pack(side="left", padx=(15, 5))
        self.sheet_combo = ttk.Combobox(top_bar, textvariable=self.sheet_var, state="readonly", width=20)
        self.sheet_combo.pack(side="left")
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda e: self.load_current_sheet())

        ttk.Label(top_bar, text="💡 Click to speak.").pack(side="right", padx=10)

        # ==========================================
        # 2. 左侧配置边栏 (抽屉式)
        # ==========================================
        self.side_bar = ttk.Frame(self.root, padding=(10, 5, 10, 10), width=280)
        self.side_bar.grid(row=1, column=0, sticky="ns")
        self.side_bar.pack_propagate(False)

        # --- Voice Settings ---
        voice_frame = ttk.LabelFrame(self.side_bar, text="Voice Control", padding=10)
        voice_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(voice_frame, text="macOS Voice:").grid(row=0, column=0, sticky="w", pady=2)
        ttk.Entry(voice_frame, textvariable=self.voice_var, width=15).grid(row=0, column=1, sticky="ew", pady=2)
        ttk.Label(voice_frame, text="Speed:").grid(row=1, column=0, sticky="w", pady=2)
        ttk.Spinbox(voice_frame, from_=120, to=320, textvariable=self.rate_var, width=13).grid(row=1, column=1, sticky="ew", pady=2)

        # --- English Column Settings ---
        eng_frame = ttk.LabelFrame(self.side_bar, text="🇬🇧 English Column", padding=10)
        eng_frame.pack(fill="x", pady=(0, 10))
        self._build_col_settings(eng_frame, self.eng_col_var, self.eng_font_var, self.eng_size_var, self.eng_width_var, self.eng_color_var, "eng")

        # --- Chinese Column Settings ---
        chn_frame = ttk.LabelFrame(self.side_bar, text="🇨🇳 Chinese Column", padding=10)
        chn_frame.pack(fill="x", pady=(0, 10))
        self._build_col_settings(chn_frame, self.chn_col_var, self.chn_font_var, self.chn_size_var, self.chn_width_var, self.chn_color_var, "chn")

        # --- View Control ---
        view_frame = ttk.LabelFrame(self.side_bar, text="Pagination", padding=10)
        view_frame.pack(fill="x", pady=(0, 15))
        ttk.Label(view_frame, text="Rows per page:").pack(side="left")
        ttk.Spinbox(view_frame, from_=50, to=1000, increment=50, textvariable=self.page_size_var, width=8).pack(side="right")

        ttk.Button(self.side_bar, text="Apply", command=self.render_current_page).pack(fill="x", pady=(5, 5))
        ttk.Button(self.side_bar, text="Save Settings", command=self.save_file_settings).pack(fill="x", pady=(0, 5))

        # ==========================================
        # 3. 主数据区 (Canvas)
        # ==========================================
        main_frame = ttk.Frame(self.root)
        main_frame.grid(row=1, column=1, sticky="nsew", padx=(0, 10))
        main_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)

        self.table_canvas = tk.Canvas(main_frame, bg=self.theme["bg"], highlightthickness=0)
        self.table_canvas.grid(row=0, column=0, sticky="nsew")

        yscroll = ttk.Scrollbar(main_frame, orient="vertical", command=self.table_canvas.yview, style="Vertical.TScrollbar")
        yscroll.grid(row=0, column=1, sticky="ns")
        self.table_canvas.configure(yscrollcommand=yscroll.set)

        xscroll = ttk.Scrollbar(main_frame, orient="horizontal", command=self.table_canvas.xview, style="Horizontal.TScrollbar")
        xscroll.grid(row=1, column=0, sticky="ew")
        self.table_canvas.configure(xscrollcommand=xscroll.set)

        self.table_inner = tk.Frame(self.table_canvas, bg=self.theme["grid"])
        self.table_canvas.create_window((0, 0), window=self.table_inner, anchor="nw")
        self.table_inner.bind("<Configure>", lambda e: self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all")))
        
        self.table_canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.root.bind("<Down>", self.move_selection_down)
        self.root.bind("<Up>", self.move_selection_up)

        # ==========================================
        # 4. 底部状态栏
        # ==========================================
        bottom_bar = ttk.Frame(self.root, padding=8)
        bottom_bar.grid(row=2, column=0, columnspan=2, sticky="ew")
        
        page_frame = ttk.Frame(bottom_bar)
        page_frame.pack(side="right")
        self.prev_btn = ttk.Button(page_frame, text="◀ Prev", command=self.prev_page, state="disabled")
        self.prev_btn.pack(side="left", padx=5)
        ttk.Label(page_frame, textvariable=self.page_info_var).pack(side="left", padx=10)
        self.next_btn = ttk.Button(page_frame, text="Next ▶", command=self.next_page, state="disabled")
        self.next_btn.pack(side="left", padx=5)

        self.status_label = ttk.Label(bottom_bar, textvariable=self.status_var, foreground=self.theme["muted"])
        self.status_label.pack(side="left", fill="x")

    def _build_col_settings(self, parent, col_var, font_var, size_var, width_var, color_var, btn_key):
        """辅助方法：构建单独列的设置项"""
        # Column Index
        ttk.Label(parent, text="Col (e.g. A):").grid(row=0, column=0, sticky="w", pady=2)
        ttk.Entry(parent, textvariable=col_var, width=15).grid(row=0, column=1, sticky="ew", pady=2)
        # Font Selection
        ttk.Label(parent, text="Font:").grid(row=1, column=0, sticky="w", pady=2)
        ttk.Combobox(parent, textvariable=font_var, values=self.font_options, state="readonly", width=13).grid(row=1, column=1, sticky="ew", pady=2)
        # Font Size
        ttk.Label(parent, text="Size:").grid(row=2, column=0, sticky="w", pady=2)
        ttk.Spinbox(parent, from_=10, to=36, textvariable=size_var, width=13).grid(row=2, column=1, sticky="ew", pady=2)
        # Column Width
        ttk.Label(parent, text="Width (px):").grid(row=3, column=0, sticky="w", pady=2)
        ttk.Spinbox(parent, from_=100, to=1000, increment=50, textvariable=width_var, width=13).grid(row=3, column=1, sticky="ew", pady=2)
        # Text Color
        ttk.Label(parent, text="Color:").grid(row=4, column=0, sticky="w", pady=2)
        color_btn = ttk.Button(
            parent,
            text=color_var.get(),
            command=lambda: self.choose_color(color_var, btn_key),
            style="Picker.TButton",
        )
        color_btn.grid(row=4, column=1, sticky="ew", pady=2)
        self.color_buttons[btn_key] = color_btn

    def update_color_button_text(self, btn_key, color_value):
        btn = self.color_buttons.get(btn_key)
        if btn is not None:
            btn.configure(text=color_value)

    def _update_theme_button_text(self):
        if hasattr(self, "theme_btn"):
            self.theme_btn.configure(text=f"Theme: {self.theme_mode.capitalize()}")

    def apply_theme_mode(self, mode):
        if mode not in self.THEMES:
            return

        selected_row = self.current_highlight_row
        selected_col = self.current_selected_col

        self.theme_mode = mode
        self.theme = dict(self.THEMES[self.theme_mode])
        self.root.configure(bg=self.theme["bg"])
        self._configure_theme()
        self._update_theme_button_text()

        if hasattr(self, "table_canvas"):
            self.table_canvas.configure(bg=self.theme["bg"])
        if hasattr(self, "table_inner"):
            self.table_inner.configure(bg=self.theme["grid"])
        if hasattr(self, "status_label"):
            self.status_label.configure(foreground=self.theme["muted"])

        if self.workbook is not None:
            self.render_current_page()
            if selected_row is not None and selected_col is not None:
                self.current_selected_col = selected_col
                self.set_highlight_row(selected_row)
                target_widget = self.cell_labels.get((selected_row, selected_col))
                if target_widget is not None:
                    self.center_widget_in_view(target_widget)

        self.save_app_settings(quiet=True)
        self.status_var.set(f"Theme switched to {self.theme_mode}.")

    def toggle_theme(self):
        self.apply_theme_mode("light" if self.theme_mode == "dark" else "dark")

    def choose_color(self, color_var, btn_key):
        color_code = colorchooser.askcolor(title="Choose Text Color", initialcolor=color_var.get())[1]
        if color_code:
            color_var.set(color_code)
            self.update_color_button_text(btn_key, color_code)

    def save_file_settings(self):
        if not self.file_path:
            messagebox.showinfo("Save Settings", "Open an Excel file first.")
            return

        store = self._ensure_file_store()
        if store is None:
            return

        store["__ui_settings__"] = {
            key: getattr(self, var_name).get()
            for var_name, key in self.UI_SETTING_FIELDS
        }
        self.save_config(quiet=True)
        self.status_var.set("Settings saved for current file.")

    def load_file_settings(self):
        settings = self.get_file_store().get("__ui_settings__")
        if not isinstance(settings, dict):
            return

        for var_name, key in self.UI_SETTING_FIELDS:
            getattr(self, var_name).set(settings.get(key, getattr(self, var_name).get()))

        self.update_color_button_text("eng", self.eng_color_var.get())
        self.update_color_button_text("chn", self.chn_color_var.get())

    def toggle_settings(self):
        if self.settings_visible:
            self.side_bar.grid_remove()
        else:
            self.side_bar.grid()
        self.settings_visible = not self.settings_visible

    def _on_mousewheel(self, event):
        self.table_canvas.yview_scroll(int(-1 * event.delta), "units")

    def make_bookmark_key(self, row, col):
        return f"{row}:{col}"

    def parse_bookmark_key(self, key):
        row, col = key.split(":")
        return int(row), int(col)

    def get_sheet_bookmarks(self):
        if not self.file_path:
            return set()
        data = self.get_file_store().get(self.sheet_var.get(), [])
        if isinstance(data, int):
            return {self.make_bookmark_key(data, 1)}
        if isinstance(data, list):
            return {x for x in data if isinstance(x, str) and ":" in x}
        return set()

    def save_sheet_bookmarks(self, bookmarks):
        store = self._ensure_file_store()
        if store is None:
            return
        ordered = sorted(bookmarks, key=lambda x: self.parse_bookmark_key(x))
        store[self.sheet_var.get()] = ordered
        self.save_config()

    def cancel_pending_click(self):
        if self.single_click_job is not None:
            self.root.after_cancel(self.single_click_job)
            self.single_click_job = None
        self.pending_click_args = None

    def schedule_cell_click(self, excel_row, excel_col, text, target_widget):
        self.cancel_pending_click()
        self.pending_click_args = (excel_row, excel_col, text, target_widget)
        self.single_click_job = self.root.after(220, self.run_pending_click)

    def run_pending_click(self):
        self.single_click_job = None
        if self.pending_click_args:
            args = self.pending_click_args
            self.pending_click_args = None
            self.on_cell_click(*args)

    def set_cell_bg(self, widget, bg):
        widget.config(bg=bg)
        for child in widget.winfo_children():
            try:
                child.configure(bg=bg)
            except tk.TclError:
                pass

    def set_highlight_row(self, excel_row):
        if self.current_highlight_row and self.current_highlight_row in self.row_labels:
            for lbl in self.row_labels[self.current_highlight_row]:
                self.set_cell_bg(lbl, self.theme["cell"])

        self.current_highlight_row = excel_row
        if excel_row in self.row_labels:
            for lbl in self.row_labels[excel_row]:
                self.set_cell_bg(lbl, self.theme["accent_soft"])

    def add_bookmark_badge(self, target_widget):
        self.remove_bookmark_badge(target_widget)
        badge = tk.Label(
            target_widget,
            text="♦️",
            bg=target_widget.cget("bg"),
            bd=0,
            padx=0,
            pady=0,
            fg=self.theme["accent"],
            font=("TkDefaultFont", 11),
        )
        badge.place(relx=1.0, rely=0.0, x=-4, y=4, anchor="ne")
        badge.bind("<Double-Button-1>", lambda e, w=target_widget: self.on_badge_double_click(w))
        badge.lift()
        target_widget._bookmark_badge = badge

    def remove_bookmark_badge(self, target_widget):
        badge = getattr(target_widget, "_bookmark_badge", None)
        if badge is not None and badge.winfo_exists():
            badge.destroy()
        if hasattr(target_widget, "_bookmark_badge"):
            delattr(target_widget, "_bookmark_badge")

    def on_badge_double_click(self, target_widget):
        excel_row = getattr(target_widget, "_excel_row", None)
        excel_col = getattr(target_widget, "_excel_col", None)
        if excel_row is None or excel_col is None:
            return "break"
        return self.on_cell_double_click(excel_row, excel_col, target_widget)

    def clear_table(self):
        for widget in self.table_inner.winfo_children():
            widget.destroy()
        self.row_labels.clear()
        self.cell_labels.clear()
        self.current_highlight_row = None

    def is_target_col(self, c_idx, col_str):
        if not col_str: return False
        try:
            if col_str.isdigit(): return int(col_str) == c_idx
            return column_index_from_string(col_str.upper()) == c_idx
        except:
            return False

    def format_spaced_multiline_text(self, text, font, max_width):
        s = "" if text is None else str(text)
        if not s:
            return s

        f = tkfont.Font(font=font)
        wrapped_lines = []

        for raw_line in s.splitlines() or [""]:
            if raw_line == "":
                wrapped_lines.append("")
                continue

            current = ""
            for ch in raw_line:
                trial = current + ch
                if current and f.measure(trial) > max_width:
                    wrapped_lines.append(current)
                    current = ch
                else:
                    current = trial
            if current:
                wrapped_lines.append(current)

        return "\n\n".join(wrapped_lines)

    def create_cell(self, text, r_idx, c_idx, is_header=False, excel_row=None, excel_col=None):
        bg_color = self.theme["header"] if is_header else self.theme["cell"]
        fg_color = self.theme["muted"]
        font_name = "TkDefaultFont"
        font_size = 10
        width_px = 100 # 默认宽度

        if is_header:
            # 动态决定表头宽度
            if self.is_target_col(c_idx, self.eng_col_var.get()): width_px = self.eng_width_var.get()
            elif self.is_target_col(c_idx, self.chn_col_var.get()): width_px = self.chn_width_var.get()
            elif c_idx == 0: width_px = 60 # 序号列
        else:
            # 核心：根据列号应用独立样式
            if self.is_target_col(c_idx, self.eng_col_var.get()):
                font_name = self.eng_font_var.get()
                font_size = self.eng_size_var.get()
                fg_color = self.eng_color_var.get()
                width_px = self.eng_width_var.get()
            elif self.is_target_col(c_idx, self.chn_col_var.get()):
                font_name = self.chn_font_var.get()
                font_size = self.chn_size_var.get()
                fg_color = self.chn_color_var.get()
                width_px = self.chn_width_var.get()
            else:
                fg_color = self.theme["subtle"]

        # 处理系统默认字体回退
        if font_name == "System Default" or not font_name:
            font = ("TkDefaultFont", font_size)
        else:
            font = (font_name, font_size)

        display_text = text
        if (not is_header) and self.is_target_col(c_idx, self.chn_col_var.get()):
            display_text = self.format_spaced_multiline_text(text, font, width_px - 20)

        lbl = tk.Label(
            self.table_inner,
            text=display_text,
            bg=bg_color,
            fg=fg_color,
            font=font,
            relief="flat",
            bd=0,
            highlightthickness=0,
            activebackground=bg_color,
            activeforeground=fg_color,
            anchor="nw", justify="left", padx=10, pady=10, wraplength=width_px - 20
        )
        lbl.grid(row=r_idx, column=c_idx, sticky="nsew", padx=(0, 1), pady=(0, 1))
        lbl.bind("<MouseWheel>", self._on_mousewheel)

        # 只记录需要交互的列，防止无用列干扰
        if not is_header and excel_row is not None:
            lbl._excel_row = excel_row
            lbl._excel_col = excel_col
            lbl._cell_text = text
            if excel_row not in self.row_labels:
                self.row_labels[excel_row] = []
            self.row_labels[excel_row].append(lbl)
            self.cell_labels[(excel_row, excel_col)] = lbl
            lbl.bind("<Button-1>", lambda e, er=excel_row, ec=excel_col, txt=text, widget=lbl: self.schedule_cell_click(er, ec, txt, widget))
            lbl.bind("<Double-Button-1>", lambda e, er=excel_row, ec=excel_col, widget=lbl: self.on_cell_double_click(er, ec, widget))
        
        return width_px

    def open_file(self):
        if self.workbook is not None and self.file_path:
            self.save_resume_position()

        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")])
        if not path: return
        try:
            self.stop_speaking()
            self.workbook = load_workbook(path, data_only=True)
            self.file_path = path
            self.load_file_settings()
            self.sheet_combo["values"] = self.workbook.sheetnames
            self.sheet_var.set(self.get_saved_sheet_name() or self.workbook.sheetnames[0])
            self.load_current_sheet()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open.\n{e}")

    def load_current_sheet(self):
        if self.workbook is None: return

        page_size = max(1, int(self.page_size_var.get()))
        saved_pos = self.get_saved_position()

        if saved_pos:
            self.current_page = (saved_pos["row"] - 1) // page_size
        else:
            bookmarks = self.get_sheet_bookmarks()
            if bookmarks:
                first_row = min(self.parse_bookmark_key(x)[0] for x in bookmarks)
                self.current_page = (first_row - 1) // page_size
            else:
                self.current_page = 0
            
        self.current_selected_col = None
        self.last_spoken_cell_info = None
        self.render_current_page()

        if saved_pos:
            ws = self.workbook[self.sheet_var.get()]
            saved_row = min(saved_pos["row"], ws.max_row or 1)
            saved_col = min(saved_pos["col"], ws.max_column or 1)
            self.current_selected_col = saved_col
            self.set_highlight_row(saved_row)
            target_widget = self.cell_labels.get((saved_row, saved_col))
            if target_widget is not None:
                self.center_widget_in_view(target_widget)

    @staticmethod
    def cell_to_text(value):
        if value is None: return ""
        if isinstance(value, datetime): return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date): return value.strftime("%Y-%m-%d")
        if isinstance(value, time): return value.strftime("%H:%M:%S")
        return str(value)

    def render_current_page(self):
        if self.workbook is None: return
        ws = self.workbook[self.sheet_var.get()]
        page_size = max(1, int(self.page_size_var.get()))

        real_rows = ws.max_row or 1
        real_cols = ws.max_column or 1
        self.total_pages = max(1, (real_rows + page_size - 1) // page_size)
        self.current_page = max(0, min(self.current_page, self.total_pages - 1))

        start_row = self.current_page * page_size + 1
        end_row = min(real_rows, start_row + page_size - 1)

        self.clear_table()
        self.table_inner.configure(bg=self.theme["grid"])

        # 1. 渲染表头并配置列宽
        self.table_inner.grid_columnconfigure(0, minsize=60)
        self.create_cell("#", 0, 0, is_header=True)
        for c in range(1, real_cols + 1):
            w = self.create_cell(get_column_letter(c), 0, c, is_header=True)
            self.table_inner.grid_columnconfigure(c, minsize=w)

        # 2. 渲染数据
        grid_row = 1
        for r in range(start_row, end_row + 1):
            self.create_cell(str(r), grid_row, 0, is_header=True)
            for c in range(1, real_cols + 1):
                text = self.cell_to_text(ws.cell(row=r, column=c).value)
                self.create_cell(text, grid_row, c, is_header=False, excel_row=r, excel_col=c)
            grid_row += 1

        self.table_canvas.yview_moveto(0)
        self.table_canvas.xview_moveto(0)

        self.page_info_var.set(f"Page {self.current_page + 1} of {self.total_pages}")
        self.prev_btn.state(["!disabled"] if self.current_page > 0 else ["disabled"])
        self.next_btn.state(["!disabled"] if self.current_page < self.total_pages - 1 else ["disabled"])
        self.status_var.set(f"Loaded | Rows {start_row}-{end_row} / {real_rows}")

        bookmarks = self.get_sheet_bookmarks()
        for key in bookmarks:
            row, col = self.parse_bookmark_key(key)
            if start_row <= row <= end_row:
                target_widget = self.cell_labels.get((row, col))
                if target_widget is not None:
                    self.add_bookmark_badge(target_widget)

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.render_current_page()
            self.save_resume_position()

    def next_page(self):
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.render_current_page()
            self.save_resume_position()

    def center_widget_in_view(self, widget):
        if widget is None or not widget.winfo_exists():
            return

        self.table_canvas.update_idletasks()
        self.table_inner.update_idletasks()

        scroll_bbox = self.table_canvas.bbox("all")
        if not scroll_bbox:
            return

        total_height = max(1, scroll_bbox[3] - scroll_bbox[1])
        visible_height = max(1, self.table_canvas.winfo_height())

        widget_center = widget.winfo_y() + widget.winfo_height() / 2
        target_top = widget_center - visible_height / 2

        max_top = max(0, total_height - visible_height)
        target_top = max(0, min(target_top, max_top))

        self.table_canvas.yview_moveto(target_top / total_height)

    def _move_selection(self, delta: int):
        if self.workbook is None or self.current_highlight_row is None or self.current_selected_col is None:
            return "break"

        self.cancel_pending_click()

        ws = self.workbook[self.sheet_var.get()]
        target_row = self.current_highlight_row + delta
        max_row = ws.max_row or 1
        if target_row < 1 or target_row > max_row:
            return "break"

        page_size = max(1, int(self.page_size_var.get()))
        target_page = (target_row - 1) // page_size
        if target_page != self.current_page:
            self.current_page = target_page
            self.render_current_page()

        target_col = self.current_selected_col
        target_widget = self.cell_labels.get((target_row, target_col))
        if target_widget is None:
            return "break"

        self.center_widget_in_view(target_widget)

        self.on_cell_click(
            target_row,
            target_col,
            target_widget._cell_text,
            target_widget,
        )
        return "break"

    def move_selection_down(self, event=None):
        return self._move_selection(1)

    def move_selection_up(self, event=None):
        return self._move_selection(-1)

    def on_cell_click(self, excel_row: int, excel_col: int, text: str, target_widget):
        self.set_highlight_row(excel_row)
        self.current_selected_col = excel_col

        coord = f"{get_column_letter(excel_col)}{excel_row}"
        self.last_spoken_cell_info = (coord, text, excel_col, target_widget)
        self.save_resume_position()
        self.trigger_speech(coord, text, col_idx=excel_col)

    def on_cell_double_click(self, excel_row: int, excel_col: int, target_widget):
        self.cancel_pending_click()
        self.set_highlight_row(excel_row)
        self.current_selected_col = excel_col

        key = self.make_bookmark_key(excel_row, excel_col)
        bookmarks = self.get_sheet_bookmarks()
        coord = f"{get_column_letter(excel_col)}{excel_row}"

        if key in bookmarks:
            bookmarks.remove(key)
            self.remove_bookmark_badge(target_widget)
            self.status_var.set(f"Bookmark removed [{coord}]")
        else:
            bookmarks.add(key)
            self.add_bookmark_badge(target_widget)
            self.status_var.set(f"Bookmarked [{coord}]")

        self.save_sheet_bookmarks(bookmarks)
        return "break"

    def trigger_speech(self, coord, text, col_idx):
        if not text.strip(): return
        
        # 只朗读配置为英文的列
        if self.is_target_col(col_idx, self.eng_col_var.get()):
            self.status_var.set(f"🗣 Speaking [{coord}]")
            self.speak_text(text)
        else:
            self.status_var.set(f"⚠️ [{coord}] skipped (Not the English column).")

    def speak_text(self, text: str):
        clean = " ".join(text.strip().split())
        if not clean: return
        self.stop_speaking()
        try:
            self.say_proc = subprocess.Popen(
                ["say", "-v", self.voice_var.get() or "Samantha", "-r", str(self.rate_var.get()), clean],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
        except Exception as e:
            messagebox.showerror("Speech error", f"Could not start speech.\n\n{e}")

    def stop_speaking(self):
        if self.say_proc and self.say_proc.poll() is None:
            try: self.say_proc.terminate()
            except: pass
        self.say_proc = None

    def on_close(self):
        self.cancel_pending_click()
        self.stop_speaking()
        self.save_resume_position()
        self.root.destroy()

if __name__ == "__main__":
    app = ExcelSpeakerApp(tk.Tk())
    app.root.mainloop()
