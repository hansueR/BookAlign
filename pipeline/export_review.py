from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .align_dp import AlignmentRow
from .config import EXCEL_COLUMNS


_STATUS_FILLS = {
    "auto_good": "E8F5E9",
    "needs_review": "FFF8E1",
    "bad_suspect": "FFEBEE",
    "skip_en": "E3F2FD",
    "skip_zh": "F3E5F5",
}


def export_aligned_xlsx(rows: Iterable[AlignmentRow], out_path: str | Path) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "aligned"
    ws.append(EXCEL_COLUMNS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in rows:
        values = [getattr(row, col) for col in EXCEL_COLUMNS]
        ws.append(values)
        excel_row = ws.max_row
        fill_color = _STATUS_FILLS.get(row.status)
        if fill_color:
            fill = PatternFill("solid", fgColor=fill_color)
            for cell in ws[excel_row]:
                cell.fill = fill
        for cell in ws[excel_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 70,
        "B": 55,
        "C": 10,
        "D": 16,
        "E": 18,
        "F": 24,
        "G": 36,
        "H": 36,
        "I": 12,
        "J": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{ws.max_row}"

    wb.save(out_path)
    return out_path
